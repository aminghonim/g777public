"""
Excel Processor - Antigravity Suite
معالج ملفات Excel لقراءة جهات الاتصال
"""

import pandas as pd
import logging
from pathlib import Path
from typing import List, Dict, Optional, Any

logger = logging.getLogger(__name__)




class ExcelProcessor:
    """معالج ملفات Excel"""
    
    def __init__(self):
        self.current_file = None
        self.current_df = None
    
    def get_sheets(self, excel_path: str) -> List[str]:
        """الحصول على قائمة الأوراق في ملف Excel (نسخة مطورة ومستقرة)"""
        try:
            # استخدام pandas ExcelFile هو الحل الأكثر استقراراً لقراءة أسماء الأوراق
            xl = pd.ExcelFile(excel_path, engine='openpyxl')
            return xl.sheet_names
        except Exception as e:
            logger.error(f"فشل قراءة الأوراق عبر pandas: {e}")
            # Fallback لـ openpyxl اليدوي في حالة الضرورة القصوى
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
                sheets = list(wb.sheetnames)
                wb.close()
                return sheets
            except Exception as ex:
                logger.critical(f"فشل كلي في قراءة هيكل الملف: {ex}")
                return []
    
    def get_columns(self, excel_path: str, sheet_name: Any = 0) -> List[str]:
        """الحصول على أسماء الأعمدة"""
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, nrows=0, engine='openpyxl')
            return list(df.columns)
        except Exception as e:
            logger.error(f"فشل قراءة الأعمدة: {e}")
            return []
    
    def read_contacts(self, excel_path: str, sheet_name: Any = 0, 
                      phone_column: Optional[str] = None, name_column: Optional[str] = None,
                      row_from: int = 0, row_to: Optional[int] = None) -> List[Dict]:
        """
        قراءة جهات الاتصال من Excel
        
        Args:
            excel_path: مسار ملف Excel
            sheet_name: اسم الورقة
            phone_column: اسم عمود الأرقام (اختياري - سيتم الكشف التلقائي)
            name_column: اسم عمود الأسماء (اختياري)
            row_from: البداية من صف
            row_to: النهاية عند صف
        
        Returns:
            قائمة من القواميس {phone, name}
        """
        try:
            # قراءة الملف
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                dtype=str,
                engine='openpyxl'
            )
            
            if df.empty:
                logger.warning("الملف فارغ!")
                return []
            
            # تحديد الصفوف
            if row_to is not None:
                df = df.iloc[row_from:row_to]
            else:
                df = df.iloc[row_from:]
            
            # الكشف التلقائي عن الأعمدة
            cols = [str(c).lower() for c in df.columns]
            
            # البحث عن عمود الأرقام
            phone_col_idx = 0
            if phone_column:
                try:
                    phone_col_idx = list(df.columns).index(phone_column)
                except:
                    pass
            else:
                for i, col in enumerate(cols):
                    if any(k in col for k in ['phone', 'number', 'mobile', 'الأرقام', 'رقم', 'هاتف', 'موبايل']):
                        phone_col_idx = i
                        break
            
            # البحث عن عمود الأسماء
            name_col_idx = None
            if name_column:
                try:
                    name_col_idx = list(df.columns).index(name_column)
                except:
                    pass
            else:
                for i, col in enumerate(cols):
                    if any(k in col for k in ['name', 'الاسم', 'اسم', 'العميل', 'customer']):
                        name_col_idx = i
                        break
            
            # استخراج البيانات
            contacts = []
            for _, row in df.iterrows():
                phone = str(row.iloc[phone_col_idx]).strip()
                name = str(row.iloc[name_col_idx]).strip() if name_col_idx is not None else 'العميل'
                
                # تنظيف الرقم
                phone = self._clean_phone(phone)
                
                if phone and phone != 'nan':
                    contacts.append({
                        'phone': phone,
                        'name': name if name and name != 'nan' else 'العميل'
                    })
            
            logger.info(f"تم تحميل {len(contacts)} جهة اتصال")
            return contacts
            
        except Exception as e:
            logger.error(f"فشل قراءة الملف: {e}")
            return []
    
    def _clean_phone(self, phone: str) -> str:
        """تنظيف رقم الهاتف"""
        if not phone or str(phone).lower() in ['nan', 'none', 'null']:
            return ''
        
        # 1. Extract digits only first
        clean_digits = ''.join(filter(str.isdigit, str(phone)))
        
        # If no digits or too short, return empty (invalid)
        if len(clean_digits) < 5: # Minimal length check
            return ''

        # 2. Logic based on digits
        final_phone = ""
        
        # Check for Egyptian pattern
        # Case: 01xxxxxxxxx (11 digits, starts with 01)
        if clean_digits.startswith('01') and len(clean_digits) == 11:
            final_phone = '+2' + clean_digits

        # Case: 01xxxxxxxx (10 digits) - Handle as Egyptian with +2 prefix
        elif clean_digits.startswith('01') and len(clean_digits) == 10:
            final_phone = '+2' + clean_digits
            
        # Case: 1xxxxxxxxx (10 digits, starts with 1) - likely missing 0
        elif clean_digits.startswith('1') and len(clean_digits) == 10:
            final_phone = '+20' + clean_digits
            
        # Case: 201xxxxxxxxx (12 digits, starts with 201) - already has country code partial
        elif clean_digits.startswith('201') and len(clean_digits) == 12:
            final_phone = '+' + clean_digits
            
        # Case: 00201... (starts with 00)
        elif clean_digits.startswith('00'):
             final_phone = '+' + clean_digits[2:]
             
        # General case
        else:
             # Just prepend + if meant to be international
             final_phone = '+' + clean_digits

        return final_phone
    
    def preview_data(self, excel_path: str, sheet_name: Any = 0, rows: int = 5) -> pd.DataFrame:
        """معاينة البيانات"""
        try:
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                nrows=rows,
                engine='openpyxl'
            )
            return df
        except Exception as e:
            logger.error(f"فشل المعاينة: {e}")
            return pd.DataFrame()
    
    def count_rows(self, excel_path: str, sheet_name: Any = 0) -> int:
        """عدد الصفوف في الورقة"""
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')
            return len(df)
        except:
            return 0
